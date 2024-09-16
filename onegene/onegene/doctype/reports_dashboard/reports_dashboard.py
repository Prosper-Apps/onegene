import frappe
from frappe.utils import add_days
from frappe.utils.csvutils import UnicodeWriter
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from six import BytesIO
from datetime import date, timedelta, datetime,time

class ReportsDashboard(Document):
	pass

@frappe.whitelist()
def download():
	filename = 'Manpower_Plan_vs_Actual_Report.xlsx'
	xlsx_content = build_xlsx_content()
	frappe.response['filename'] = filename
	frappe.response['filecontent'] = xlsx_content
	frappe.response['type'] = 'binary'

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	if wb is None:
		wb = openpyxl.Workbook()
	ws = wb.create_sheet(sheet_name, 0)

	header_date = title(args)
	ws.append(header_date)

	header_date = title1(args)
	ws.append(header_date)

	header_date = title2(args)
	ws.append(header_date)

	header_date = title3(args)
	ws.append(header_date)

	data_rows = title4(args)
	for row in data_rows:
		ws.append(row)

	header_date = title5(args)
	ws.append(header_date)

	data_rows = title6(args)
	for row in data_rows:
		ws.append(row)

	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column= 14)
	ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column= 14)
	ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column= 1)
	ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column= 2)
	ws.merge_cells(start_row=2, start_column=3, end_row=3, end_column= 3)
	ws.merge_cells(start_row=2, start_column=4, end_row=3, end_column= 4)
	ws.merge_cells(start_row=2, start_column=5, end_row=3, end_column= 5)
	ws.merge_cells(start_row=2, start_column=14, end_row=3, end_column= 14)
	ws.merge_cells(start_row=2, start_column=6, end_row=2, end_column=9 )
	ws.merge_cells(start_row=2, start_column=10, end_row=2, end_column= 13)
	# ws.merge_cells(start_row=2, start_column=14, end_row=2, end_column=16)
	ws.merge_cells(start_row=4+len(title4(args))+1, start_column=1, end_row=4+len(title4(args))+1, end_column=14)
	ws.merge_cells(start_row=4+len(title4(args)), start_column=1, end_row=4+len(title4(args)), end_column=3)
	ws.merge_cells(start_row=4+len(title4(args))-1, start_column=1, end_row=4+len(title4(args))-1, end_column=3)
	ws.merge_cells(start_row=(len(title4(args))+ len(title6(args)) + 2), start_column=1, end_row=(len(title4(args))+ len(title6(args)) + 2), end_column=3)
	ws.merge_cells(start_row=(len(title4(args))+ len(title6(args)) + 5), start_column=1, end_row=(len(title4(args))+ len(title6(args)) + 5), end_column=3)
	ws.merge_cells(start_row=(len(title4(args))+ len(title6(args)) + 4), start_column=1, end_row=(len(title4(args))+ len(title6(args)) + 4), end_column=3)
	ws.merge_cells(start_row=(len(title4(args))+ len(title6(args)) + 3), start_column=1, end_row=(len(title4(args))+ len(title6(args)) + 3), end_column=3)
	align_center = Alignment(horizontal='center',vertical='center')
	border = Border(
		left=Side(border_style='thin'),
		right=Side(border_style='thin'),
		top=Side(border_style='thin'),
		bottom=Side(border_style='thin'))
	for rows in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=17):
		for cell in rows:
			cell.font = Font(bold=True)
			cell.alignment = align_center
	for rows in ws.iter_rows(min_row=4+len(title4(args))-1, max_row=4+len(title4(args))+1, min_col=1, max_col=14):
		for cell in rows:
			cell.font = Font(bold=True)
			cell.alignment = align_center
	for rows in ws.iter_rows(min_row=(len(title4(args))+ len(title6(args)) + 2), max_row=(len(title4(args))+ len(title6(args)) + 5), min_col=1, max_col=14):
		for cell in rows:
			cell.font = Font(bold=True)
			cell.alignment = align_center
	for rows in ws.iter_rows(min_row=5, max_row=(len(title4(args))+ len(title6(args)) + 5), min_col=4, max_col=14):
		for cell in rows:
			cell.alignment = align_center
	for rows in ws.iter_rows(min_row=1, max_row=(len(title4(args))+ len(title6(args)) + 5), min_col=1, max_col=14):
		for cell in rows:
			cell.border = border
	ws.freeze_panes = 'D4'
	ws.column_dimensions['N'].width = 25
	ws.column_dimensions['B'].width = 25
	ws.column_dimensions['C'].width = 20
	ws.column_dimensions['D'].width = 20
	ws.column_dimensions['E'].width = 20
	ws.column_dimensions['Q'].width = 20
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file.getvalue()

def build_xlsx_content():
	return make_xlsx(None, "Sheet1")

@frappe.whitelist()
def title(args):
	month = datetime.strptime(str(args.date),'%Y-%m-%d')
	mon = str(month.strftime('%B') +' '+ str(month.strftime('%Y')))
	data = ["Manpower Plan vs Actual Report for " + str(month.day) + " " + mon]
	return data

@frappe.whitelist()
def title1(args):
	data = ["S NO" ,"Particulars","Department","Business Plan","Actual Plan","Man Days(A)","","","","Overtime(B)","","","","Total(A+B)"]
	return data

@frappe.whitelist()
def title2(args):
	data = ["","","","","","Com","App","Cont","Total","Com","App","Cont","Total"]
	return data

@frappe.whitelist()
def title3(args):
	data = ["Production Team"]
	return data

@frappe.whitelist()
def title4(args):
	data = []
	row = []
	departments = frappe.get_all('Department', {'parent_department': 'Manufacturing - WAIP'}, order_by='name ASC')
	tbp = 0
	tacp = 0
	tco = 0
	tap = 0
	tcon = 0
	tco_ot = 0
	tap_ot = 0
	tcon_ot = 0
	tot = 0
	i = 1
	for dept in departments:
		row = [i]
		bp = frappe.get_value("Manpower Plan",{'department':dept.name,'date':args.date},['business_plan']) or 0
		acp = frappe.get_value("Manpower Plan",{'department':dept.name,'date':args.date},['plan']) or 0
		co_ma_p = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date = '%s' AND department = '%s'  AND custom_employee_category IN ("Staff","Sub Staff","Operator") AND docstatus != 2 AND status = 'Present' AND attendance_request is NULL """ % (args.date, dept.name), as_dict=True)
		co_ma_h = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date = '%s' AND department = '%s'  AND custom_employee_category IN ("Staff","Sub Staff","Operator") AND docstatus != 2 AND status = 'Half day' AND attendance_request is NULL """ % (args.date, dept.name), as_dict=True)
		count_co_ma_p1 = co_ma_p[0].get('count', 0)
		count_co_ma_h1 = co_ma_h[0].get('count', 0)
		co_ma = count_co_ma_p1 + (count_co_ma_h1 / 2)
		ap_ma_p = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date = '%s' AND department = '%s'  AND custom_employee_category IN ("Apprentice") AND docstatus != 2 AND status = 'Present' AND attendance_request is NULL """ % (args.date, dept.name), as_dict=True)
		ap_ma_h = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date = '%s' AND department = '%s'  AND custom_employee_category IN ("Apprentice") AND docstatus != 2 AND status = 'Half day' AND attendance_request is NULL """ % (args.date, dept.name), as_dict=True)
		count_ap_ma_p1 = ap_ma_p[0].get('count', 0)
		count_ap_ma_h1 = ap_ma_h[0].get('count', 0)
		ap_ma = count_ap_ma_p1 + (count_ap_ma_h1 / 2)
		con_ma_p = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date = '%s' AND department = '%s'  AND custom_employee_category IN ("Contractor") AND docstatus != 2 AND status = 'Present' AND attendance_request is NULL """ % (args.date, dept.name), as_dict=True)
		con_ma_h = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date = '%s' AND department = '%s'  AND custom_employee_category IN ("Contractor") AND docstatus != 2 AND status = 'Half day' AND attendance_request is NULL """ % (args.date, dept.name), as_dict=True)
		count_con_ma_p1 = con_ma_p[0].get('count', 0)
		count_con_ma_h1 = con_ma_h[0].get('count', 0)
		con_ma = count_con_ma_p1 + (count_con_ma_h1 / 2)
		co_ma_ot = frappe.db.sql("""SELECT sum(custom_overtime_hours) as ot FROM `tabAttendance` WHERE attendance_date = '%s' AND department = '%s'  AND custom_employee_category IN ("Staff","Sub Staff","Operator") AND docstatus != 2 AND status = 'Present'AND attendance_request is NULL """ % (args.date, dept.name), as_dict=True)
		ap_ma_ot = frappe.db.sql("""SELECT sum(custom_overtime_hours) as ot FROM `tabAttendance` WHERE attendance_date = '%s' AND department = '%s'  AND custom_employee_category IN ("Apprentice")AND docstatus != 2 AND status = 'Present'AND attendance_request is NULL """ % (args.date, dept.name), as_dict=True)
		con_ma_ot = frappe.db.sql("""SELECT sum(custom_overtime_hours) as ot FROM `tabAttendance` WHERE attendance_date = '%s' AND department = '%s'  AND custom_employee_category IN ("Contractor")AND docstatus != 2 AND status = 'Present'AND attendance_request is NULL """ % (args.date, dept.name), as_dict=True)
		co_ot_value = co_ma_ot[0].get('ot', 0)
		if co_ot_value is not None:
			count_co_ma_ot = round(co_ot_value / 8, 2)
		else:
			count_co_ma_ot = 0.0
		ap_ot_value = ap_ma_ot[0].get('ot', 0)
		if ap_ot_value is not None:
			count_ap_ma_ot = round(ap_ot_value / 8, 2)
		else:
			count_ap_ma_ot = 0.0
		con_ot_value = con_ma_ot[0].get('ot', 0)
		if con_ot_value is not None:
			count_con_ma_ot = round(con_ot_value / 8, 2)
		else:
			count_con_ma_ot = 0.0
		row += [dept.name, 'Production', bp or 0, acp or 0, co_ma,ap_ma,con_ma,(co_ma+ap_ma+con_ma),count_co_ma_ot,count_ap_ma_ot,count_con_ma_ot,(count_co_ma_ot+count_ap_ma_ot+count_con_ma_ot),((co_ma+ap_ma+con_ma)+(count_co_ma_ot+count_ap_ma_ot+count_con_ma_ot))]
		data.append(row)

		tbp += bp
		tacp += acp
		tco += co_ma
		tap += ap_ma
		tcon += con_ma 
		tco_ot += count_co_ma_ot
		tap_ot += count_ap_ma_ot
		tcon_ot += count_con_ma_ot
		tot += (co_ma+ap_ma+con_ma)+(count_co_ma_ot+count_ap_ma_ot+count_con_ma_ot)
		i += 1

	subtotal_row = ["Sub Total (A)", "", "", tbp, tacp, tco, tap, tcon, tco + tap + tcon, tco_ot, tap_ot, tcon_ot,
					tco_ot + tap_ot + tcon_ot,tot]
	data.append(subtotal_row)

	if tacp != 0:
		percentage_row = ["Percentage", "", "", "", "", "", "", "",
						  f"{((tco + tap + tcon) / tacp) * 100:.2f} %",
						  "", "", "",
						  f"{((tco_ot + tap_ot + tcon_ot) / tacp) * 100:.2f} %",
						  "", "", "",
						  f"{((tco + tap + tcon + tco_ot + tap_ot + tcon_ot) / tacp) * 100:.2f} %"]
	else:
		percentage_row = ["Percentage", "", "", "", "", "", "", "", "", "", "", "", ""]

	data.append(percentage_row)

	return data
@frappe.whitelist()
def title5(args):
	data = ["Supporting Team"]
	return data

@frappe.whitelist()
def title6(args):
	department_data = []
	department = ['Quality - WAIP', 'NPD - WAIP', 'M P L & Purchase - WAIP', 'ME -Regular - WAIP', 'Maintenance - WAIP', 'Delivery - WAIP', 'Finance - WAIP', 'HR - WAIP']

	for d in department:
		d_group = frappe.get_value("Department", {'name': d}, ['is_group'])
		if d_group == 1:
			dept = frappe.get_all("Department", {'parent_department': d}, ['*'])
			for i in dept:
				department_data.append(i.name)
		else:
			department_data.append(d)

	data = []
	row = []
	j = 1
	tbp = tacp = tco = tap = tcon = tco_ot = tap_ot = tcon_ot = tot = 0

	for i in department_data:
		bp = frappe.get_value("Manpower Plan", {'department': i, 'date': args.date}, ['business_plan']) or 0
		acp = frappe.get_value("Manpower Plan", {'department': i, 'date': args.date}, ['plan']) or 0
		
		co_ma_p = frappe.db.sql("""
			SELECT count(*) as count 
			FROM `tabAttendance` 
			WHERE 
				attendance_date = '%s' AND 
				department = '%s' AND 
				custom_employee_category IN ("Staff","Sub Staff","Operator") AND 
				docstatus != 2 AND 
				status = 'Present' AND 
				attendance_request is NULL 
		""" % (args.date, i), as_dict=True)
		
		co_ma_h = frappe.db.sql("""
			SELECT count(*) as count 
			FROM `tabAttendance` 
			WHERE 
				attendance_date = '%s' AND 
				department = '%s' AND 
				custom_employee_category IN ("Staff","Sub Staff","Operator") AND 
				docstatus != 2 AND 
				status = 'Half day' AND 
				attendance_request is NULL 
		""" % (args.date, i), as_dict=True)
		
		count_co_ma_p1 = co_ma_p[0].get('count', 0)
		count_co_ma_h1 = co_ma_h[0].get('count', 0)
		co_ma = count_co_ma_p1 + (count_co_ma_h1 / 2)
		
		ap_ma_p = frappe.db.sql("""
			SELECT count(*) as count 
			FROM `tabAttendance` 
			WHERE 
				attendance_date = '%s' AND 
				department = '%s' AND 
				custom_employee_category IN ("Apprentice")AND 
				docstatus != 2 AND 
				status = 'Present' AND 
				attendance_request is NULL 
		""" % (args.date, i), as_dict=True)
		
		ap_ma_h = frappe.db.sql("""
			SELECT count(*) as count 
			FROM `tabAttendance` 
			WHERE 
				attendance_date = '%s' AND 
				department = '%s' AND 
				custom_employee_category IN ("Apprentice")AND 
				docstatus != 2 AND 
				status = 'Half day' AND 
				attendance_request is NULL 
		""" % (args.date, i), as_dict=True)
		
		count_ap_ma_p1 = ap_ma_p[0].get('count', 0)
		count_ap_ma_h1 = ap_ma_h[0].get('count', 0)
		ap_ma = count_ap_ma_p1 + (count_ap_ma_h1 / 2)
		
		con_ma_p = frappe.db.sql("""
			SELECT count(*) as count 
			FROM `tabAttendance` 
			WHERE 
				attendance_date = '%s' AND 
				department = '%s' AND 
				custom_employee_category IN ("Contractor") AND 
				docstatus != 2 AND 
				status = 'Present' AND 
				attendance_request is NULL 
		""" % (args.date, i), as_dict=True)
		
		con_ma_h = frappe.db.sql("""
			SELECT count(*) as count 
			FROM `tabAttendance` 
			WHERE 
				attendance_date = '%s' AND 
				department = '%s' AND 
				custom_employee_category IN ("Contractor") AND 
				docstatus != 2 AND 
				status = 'Half day' AND 
				attendance_request is NULL 
		""" % (args.date, i), as_dict=True)
		
		count_con_ma_p1 = con_ma_p[0].get('count', 0)
		count_con_ma_h1 = con_ma_h[0].get('count', 0)
		con_ma = count_con_ma_p1 + (count_con_ma_h1 / 2)
		
		co_ma_ot = frappe.db.sql("""
			SELECT sum(custom_overtime_hours) as ot 
			FROM `tabAttendance` 
			WHERE 
				attendance_date = '%s' AND 
				department = '%s' AND 
				custom_employee_category IN ("Staff","Sub Staff","Operator") AND 
				docstatus != 2 AND 
				status = 'Present' AND 
				attendance_request is NULL 
		""" % (args.date, i), as_dict=True)
		
		ap_ma_ot = frappe.db.sql("""
			SELECT sum(custom_overtime_hours) as ot 
			FROM `tabAttendance` 
			WHERE 
				attendance_date = '%s' AND 
				department = '%s' AND 
				custom_employee_category IN ("Apprentice") AND 
				docstatus != 2 AND 
				status = 'Present' AND 
				attendance_request is NULL 
		""" % (args.date, i), as_dict=True)
		
		con_ma_ot = frappe.db.sql("""
			SELECT sum(custom_overtime_hours) as ot 
			FROM `tabAttendance` 
			WHERE 
				attendance_date = '%s' AND 
				department = '%s' AND 
				custom_employee_category IN ("Contractor") AND 
				docstatus != 2 AND 
				status = 'Present' AND 
				attendance_request is NULL 
		""" % (args.date, i), as_dict=True)
		
		co_ot_value = co_ma_ot[0].get('ot', 0)
		count_co_ma_ot = round(co_ot_value / 8, 2) if co_ot_value is not None else 0.0
		
		ap_ot_value = ap_ma_ot[0].get('ot', 0)
		count_ap_ma_ot = round(ap_ot_value / 8, 2) if ap_ot_value is not None else 0.0
		
		con_ot_value = con_ma_ot[0].get('ot', 0)
		count_con_ma_ot = round(con_ot_value / 8, 2) if con_ot_value is not None else 0.0
		
		row = [j, i, i, bp or 0, acp or 0, co_ma, ap_ma, con_ma, (co_ma + ap_ma + con_ma), count_co_ma_ot, count_ap_ma_ot, count_con_ma_ot, (count_co_ma_ot + count_ap_ma_ot + count_con_ma_ot), (((co_ma + ap_ma + con_ma)) + ((count_co_ma_ot + count_ap_ma_ot + count_con_ma_ot)))]
		data.append(row)
		
		j += 1
		tbp += bp
		tacp += acp
		tco += co_ma
		tap += ap_ma
		tcon += con_ma
		tco_ot += count_co_ma_ot
		tap_ot += count_ap_ma_ot
		tcon_ot += count_con_ma_ot
		tot += ((co_ma + ap_ma + con_ma)) + ((count_co_ma_ot + count_ap_ma_ot + count_con_ma_ot))

	row1 = ["Sub Total (B)", " ", " ", tbp, tacp, tco, tap, tcon, (tco + tap + tcon), tco_ot, tap_ot, tcon_ot, (tco_ot + tap_ot + tcon_ot),tot]

	if tacp != 0:
		row2 = ["Percentage", " ", " ", " ", " ", " ", " ", " ", (((tco + tap + tcon) / tacp) * 100) + ' %',(((tco_ot + tap_ot + tcon_ot) / tacp) * 100) + ' %', " ", " ", " ", (((tot) / tacp) * 100) + ' %']
	else:
		row2 = ["Percentage", " ", " ", " ", " ", " ", " ", " ", " ", " ", " ", " ", " ", " "]

	data.append(row1)
	data.append(row2)

	departments = frappe.get_all('Department', {'parent_department': 'Manufacturing - WAIP'}, order_by='name ASC')

	total_business_plan = tbp
	total_actual_plan = tacp
	total_staff_count = tco
	total_apprentice_count = tap
	total_contractor_count = tcon
	total_staff_overtime_hours = tco_ot
	total_apprentice_overtime_hours = tap_ot
	total_contractor_overtime_hours = tcon_ot

	for dept in departments:
		business_plan = frappe.get_value("Manpower Plan", {'department': dept.name, 'date': args.date}, ['business_plan']) or 0
		actual_plan = frappe.get_value("Manpower Plan", {'department': dept.name, 'date': args.date}, ['plan']) or 0
		
		staff_present_query = frappe.db.sql(f"""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date = '{args.date}' AND department = '{dept.name}'  
												AND custom_employee_category IN ("Staff","Sub Staff","Operator") AND docstatus != 2 
												AND status = 'Present' AND attendance_request is NULL """, as_dict=True)
		staff_half_day_query = frappe.db.sql(f"""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date = '{args.date}' AND department = '{dept.name}'  
												AND custom_employee_category IN ("Staff","Sub Staff","Operator") AND docstatus != 2 
												AND status = 'Half day' AND attendance_request is NULL """, as_dict=True)
		staff_present_count = staff_present_query[0].get('count', 0)
		staff_half_day_count = staff_half_day_query[0].get('count', 0)
		staff_count = staff_present_count + (staff_half_day_count / 2)

		apprentice_present_query = frappe.db.sql(f"""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date = '{args.date}' AND department = '{dept.name}'  
													AND custom_employee_category IN ("Apprentice") AND docstatus != 2 
													AND status = 'Present' AND attendance_request is NULL """, as_dict=True)
		apprentice_half_day_query = frappe.db.sql(f"""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date = '{args.date}' AND department = '{dept.name}'  
													AND custom_employee_category IN ("Apprentice") AND docstatus != 2 
													AND status = 'Half day' AND attendance_request is NULL """, as_dict=True)
		apprentice_present_count = apprentice_present_query[0].get('count', 0)
		apprentice_half_day_count = apprentice_half_day_query[0].get('count', 0)
		apprentice_count = apprentice_present_count + (apprentice_half_day_count / 2)

		contractor_present_query = frappe.db.sql(f"""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date = '{args.date}' AND department = '{dept.name}'  
													AND custom_employee_category IN ("Contractor") AND docstatus != 2 
													AND status = 'Present' AND attendance_request is NULL """, as_dict=True)
		contractor_half_day_query = frappe.db.sql(f"""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date = '{args.date}' AND department = '{dept.name}'  
													AND custom_employee_category IN ("Contractor") AND docstatus != 2 
													AND status = 'Half day' AND attendance_request is NULL """, as_dict=True)
		contractor_present_count = contractor_present_query[0].get('count', 0)
		contractor_half_day_count = contractor_half_day_query[0].get('count', 0)
		contractor_count = contractor_present_count + (contractor_half_day_count / 2)

		staff_overtime_query = frappe.db.sql(f"""SELECT sum(custom_overtime_hours) as ot FROM `tabAttendance` WHERE attendance_date = '{args.date}' AND department = '{dept.name}'  
												AND custom_employee_category IN ("Staff","Sub Staff","Operator") AND docstatus != 2 
												AND status = 'Present' AND attendance_request is NULL """, as_dict=True)
		apprentice_overtime_query = frappe.db.sql(f"""SELECT sum(custom_overtime_hours) as ot FROM `tabAttendance` WHERE attendance_date = '{args.date}' AND department = '{dept.name}'  
													AND custom_employee_category IN ("Apprentice") AND docstatus != 2 
													AND status = 'Present' AND attendance_request is NULL """, as_dict=True)
		contractor_overtime_query = frappe.db.sql(f"""SELECT sum(custom_overtime_hours) as ot FROM `tabAttendance` WHERE attendance_date = '{args.date}' AND department = '{dept.name}'  
													AND custom_employee_category IN ("Contractor") AND docstatus != 2 
													AND status = 'Present' AND attendance_request is NULL """, as_dict=True)
		staff_overtime_hours = staff_overtime_query[0].get('ot', 0) / 8 if staff_overtime_query[0].get('ot', 0) else 0
		apprentice_overtime_hours = apprentice_overtime_query[0].get('ot', 0) / 8 if apprentice_overtime_query[0].get('ot', 0) else 0
		contractor_overtime_hours = contractor_overtime_query[0].get('ot', 0) / 8 if contractor_overtime_query[0].get('ot', 0) else 0

		total_business_plan += business_plan
		total_actual_plan += actual_plan
		total_staff_count += staff_count
		total_apprentice_count += apprentice_count
		total_contractor_count += contractor_count
		total_staff_overtime_hours += staff_overtime_hours
		total_apprentice_overtime_hours += apprentice_overtime_hours
		total_contractor_overtime_hours += contractor_overtime_hours

	if total_actual_plan != 0:
		staff_percentage = ((total_staff_count) / total_actual_plan) * 100
		apprentice_percentage = ((total_apprentice_count) / total_actual_plan) * 100
		contractor_percentage = ((total_contractor_count) / total_actual_plan) * 100
	else:
		staff_percentage = 0
		apprentice_percentage = 0
		contractor_percentage = 0

	data.append(["Total (A + B)", "", "", total_business_plan, total_actual_plan, total_staff_count, total_apprentice_count, total_contractor_count, 
				(total_staff_count + total_apprentice_count + total_contractor_count), total_staff_overtime_hours, total_apprentice_overtime_hours, 
				total_contractor_overtime_hours, (total_staff_overtime_hours + total_apprentice_overtime_hours + total_contractor_overtime_hours), 
				((total_staff_count + total_apprentice_count + total_contractor_count) + 
				(total_staff_overtime_hours + total_apprentice_overtime_hours + total_contractor_overtime_hours))])

	data.append(["Percentage", "", "", "", "", "", "", "", staff_percentage, "", "", "", apprentice_percentage,  contractor_percentage])

	return data
 