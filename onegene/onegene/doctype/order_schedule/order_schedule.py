# Copyright (c) 2023, TEAMPRO and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
from frappe.utils.file_manager import get_file
from frappe.utils import cstr, add_days, date_diff, getdate
from frappe.utils.background_jobs import enqueue
import datetime
from datetime import datetime


from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import frappe
import requests
from datetime import date
import erpnext
import json
from frappe.utils import now
from frappe import throw,_
from frappe.utils import flt
from frappe.utils import (
    add_days,
    ceil,
    cint,
    comma_and,
    flt,
    get_link_to_form,
    getdate,
    now_datetime,
    datetime,get_first_day,get_last_day,
    nowdate,
    today,
)
from frappe.utils import cstr, cint, getdate, get_last_day, get_first_day, add_days,date_diff
from datetime import date, datetime, timedelta
import datetime as dt
from datetime import datetime, timedelta
from typing import Dict, Optional, Tuple, Union



class OrderSchedule(Document):
    def after_insert(self):
        if self.customer_code:
            self.customer_name = frappe.db.get_value("Customer",{"customer_code":self.customer_code},["name"])
    def on_update(self):
        if self.sales_order_number:
            order_type = frappe.db.get_value("Sales Order",self.sales_order_number,'customer_order_type')
            if order_type == "Open":
                order_open = frappe.get_doc("Open Order",{"sales_order_number":self.sales_order_number})
                for ord in order_open.open_order_table:
                    item_qty = 0
                    order_schedules = frappe.get_all("Order Schedule", {"sales_order_number": self.sales_order_number,"item_code": self.item_code}, ["qty"])
                    for sche in order_schedules:
                        item_qty += sche.qty
                    if ord.item_code == self.item_code:
                        ord.qty = item_qty
                    order_open.save(ignore_permissions=True)
            else:
                s_qty = 0
                so_exist = frappe.db.exists('Order Schedule',{"sales_order_number":self.sales_order_number,"customer_code":self.customer_code,"item_code":self.item_code})
                if so_exist:
                    exist_so = frappe.get_all("Order Schedule",{"sales_order_number":self.sales_order_number,"customer_code":self.customer_code,"item_code":self.item_code},["*"])
                    for i in exist_so:
                        old_qty = i.qty
                        s_qty += old_qty
                    if order_type == "Fixed":
                        sales_order = frappe.get_all("Sales Order Item",{"parent": self.sales_order_number, "item_code": self.item_code},["qty","item_code"])
                        if sales_order and len(sales_order) > 0:
                            sales_order_qty = sales_order[0].get("qty")
                            item_cde = sales_order[0].get("item_code")
                            idx = sales_order[0].get("idx")
                            if sales_order_qty < s_qty:
                                frappe.throw(f"Validation failed: Quantity <b>{s_qty}</b> exceeds Sales Order quantity <b>{sales_order_qty}</b> in Line Item <b>{item_cde}</b>.")
                        else:
                            frappe.throw(f"Item <b>{self.item_code}</b> not found in the Sales Order - <b>{self.sales_order_number}</b>")

@frappe.whitelist()
def return_items(doctype,docname):
    doc = frappe.get_doc(doctype,docname)
    return doc.items

@frappe.whitelist()
def qty_check(sales,item,customer,qty):
    if sales:
        order_type = frappe.db.get_value("Sales Order",sales,'customer_order_type')
        if order_type == "Fixed":
            s_qty = 0
            so_exist = frappe.db.exists('Order Schedule',{"sales_order_number":sales,"customer_code":customer,"item_code":item})
            if so_exist:
                exist_so = frappe.get_all("Order Schedule",{"sales_order_number":sales,"customer_code":customer,"item_code":item},["*"])
                for i in exist_so:
                    old_qty = i.qty
                    s_qty += old_qty
                if order_type == "Fixed":
                    sales_order = frappe.get_all("Sales Order Item",{"parent": sales, "item_code": item},["qty","item_code"])
                    if sales_order and len(sales_order) > 0:
                        total_os_qty = float(s_qty)+float(qty)
                        sales_order_qty = sales_order[0].get("qty")
                        item_cde = sales_order[0].get("item_code")
                        idx = sales_order[0].get("idx")
                        if sales_order_qty < total_os_qty:
                            frappe.throw(f"Validation failed: Quantity <b>{total_os_qty}</b> exceeds Sales Order quantity <b>{sales_order_qty}</b> in Line Item <b>{item_cde}</b>.")
                    else:
                        frappe.throw(f"Item <b>{item}</b> not found in the Sales Order - <b>{sales}</b>")

@frappe.whitelist()
def schedule_list(sales, item):
    if sales and item:
        documents = frappe.get_all('Order Schedule', {'sales_order_number': sales, 'item_code': item},
                                    ['schedule_date', 'tentative_plan_1', 'tentative_plan_2', 'qty', 'delivered_qty',
                                    'pending_qty', 'remarks', 'order_rate'])

        documents = sorted(documents, key=lambda x: x['schedule_date'])
        data = '<table border="1" style="width: 100%;">'
        data += '<tr style="background-color:#D9E2ED;">'
        data += '<td colspan="2" style="text-align:center;"><b>Schedule Month</b></td>'
        data += '<td colspan="2" style="text-align:center;"><b>Schedule Date</b></td>'
        data += '<td colspan="2" style="text-align:center;"><b>Tentative Plan - I</b></td>'
        data += '<td colspan="2" style="text-align:center;"><b>Tentative Plan - II</b></td>'
        data += '<td colspan="2" style="text-align:center;"><b>Schedule Qty</b></td>'
        data += '<td colspan="2" style="text-align:center;"><b>Delivered Qty</b></td>'
        data += '<td colspan="2" style="text-align:center;"><b>Pending Qty</b></td>'
        data += '<td colspan="2" style="text-align:center;"><b>Remarks</b></td>'
        data += '<td colspan="2" style="text-align:center;"><b>Cost Price</b></td>'
        data += '</tr>'
        for doc in documents:
            month_string = doc['schedule_date'].strftime('%B')
            data += '<tr>'
            data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(month_string)
            data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['schedule_date'].strftime('%d-%m-%Y'))
            data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['tentative_plan_1'])
            data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['tentative_plan_2'])
            data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['qty'])
            data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['delivered_qty'])
            data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['pending_qty'])
            if doc['remarks']:
                data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['remarks'])
            else:
                data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format('-')
            data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['order_rate'])
            data += '</tr>'
        data += '</table>'
        return data

@frappe.whitelist()
def get_qty_rate_so(item,sales):
    so = frappe.db.get_value("Sales Order Item",{"Parent":sales,"item_code":item},["rate"])
    return so
